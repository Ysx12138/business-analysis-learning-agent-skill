"""
生成在线零售数据分析 Excel 文件
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import os

OUT_DIR = "output/online_retail_excel"
os.makedirs(OUT_DIR, exist_ok=True)

# ── 读取清洗后数据 ──
df = pd.read_csv("/Users/sx/数据分析/online_retail/原始数据/online_retail.csv",
                 encoding="ISO-8859-1")
df["InvoiceDate"] = pd.to_datetime(df["InvoiceDate"])
df["Revenue"] = df["Quantity"] * df["UnitPrice"]

# 清洗
df = df.dropna(subset=["CustomerID"])
df = df[~df["InvoiceNo"].astype(str).str.startswith("C", na=False)]
df = df[df["Quantity"] > 0]
df = df[df["UnitPrice"] > 0]
df["Month"] = df["InvoiceDate"].dt.to_period("M").astype(str)

# ── 样式 ──
BRAND = "1B365D"
PARCHMENT = "F5F4ED"
DARK_WARM = "3D3D3A"
STONE = "6B6A64"

hdr_font = Font(name="Noto Sans SC", bold=True, color="FFFFFF", size=11)
hdr_fill = PatternFill(start_color=BRAND, end_color=BRAND, fill_type="solid")
title_font = Font(name="Noto Sans SC", bold=True, color=BRAND, size=14)
body_font = Font(name="Noto Sans SC", size=10)
num_font = Font(name="JetBrains Mono", size=10)
thin_border = Border(
    left=Side(style="thin", color="E8E6DC"),
    right=Side(style="thin", color="E8E6DC"),
    top=Side(style="thin", color="E8E6DC"),
    bottom=Side(style="thin", color="E8E6DC"),
)
kpi_fill = PatternFill(start_color="EEF2F7", end_color="EEF2F7", fill_type="solid")


def style_header(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border


def style_data(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = num_font if isinstance(cell.value, (int, float)) else body_font
            cell.alignment = Alignment(horizontal="right" if isinstance(cell.value, (int, float)) else "left")
            cell.border = thin_border


def add_title(ws, row, col, text):
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = title_font
    return cell


# ══════════════════════════════════════════
# Sheet 1: 整体对比
# ══════════════════════════════════════════
wb = Workbook()
ws1 = wb.active
ws1.title = "整体概览"
ws1.sheet_properties.tabColor = BRAND

add_title(ws1, 1, 1, "在线零售数据 · 整体概览")
ws1.merge_cells("A1:D1")

headers = ["指标", "数值", "说明", "计算口径"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 4)

metrics = [
    ("总收入", 8911407.90, "12个月线上零售总额", "sum(Revenue)"),
    ("总订单数", 18532, "有效订单数", "nunique(InvoiceNo)"),
    ("总客户数", 4338, "有购买记录的客户", "nunique(CustomerID)"),
    ("总销量(件)", 5167812, "所有商品件数之和", "sum(Quantity)"),
    ("客单价 AOV", 480.87, "平均每单金额", "Revenue / Orders"),
    ("每客户收入", 2054.27, "客户生命周期价值下限", "Revenue / Customers"),
    ("每单平均件数", 278.9, "含批发大单", "Quantity / Orders"),
    ("平均单价", 3.12, "所有商品单价均值", "mean(UnitPrice)"),
    ("复购率", "65.6%", "多次购买客户占比", "repeat / total customers"),
    ("数据保留率", "73.4%", "清洗后保留比例", "clean / raw rows"),
]

for i, (k, v, d, c) in enumerate(metrics, 4):
    ws1.cell(row=i, column=1, value=k)
    cell_v = ws1.cell(row=i, column=2, value=v)
    if isinstance(v, float):
        cell_v.number_format = '£#,##0.00' if k in ["总收入", "客单价 AOV", "每客户收入"] else '#,##0.00'
    else:
        cell_v.number_format = '#,##0'
    ws1.cell(row=i, column=3, value=d)
    ws1.cell(row=i, column=4, value=c)

style_data(ws1, 4, 3 + len(metrics), 4)
ws1.column_dimensions["A"].width = 18
ws1.column_dimensions["B"].width = 16
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 22


# ══════════════════════════════════════════
# Sheet 2: 月度趋势
# ══════════════════════════════════════════
ws2 = wb.create_sheet("月度趋势")
ws2.sheet_properties.tabColor = BRAND

add_title(ws2, 1, 1, "月度趋势")
ws2.merge_cells("A1:E1")

monthly = df.groupby("Month").agg(
    收入=("Revenue", "sum"),
    订单数=("InvoiceNo", "nunique"),
    客户数=("CustomerID", "nunique"),
    销量=("Quantity", "sum"),
)
monthly["客单价"] = monthly["收入"] / monthly["订单数"]
monthly = monthly.round(2)

headers2 = ["月份", "收入", "订单数", "客户数", "客单价"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=3, column=i, value=h)
style_header(ws2, 3, 5)

for i, (idx, row) in enumerate(monthly.iterrows(), 4):
    ws2.cell(row=i, column=1, value=str(idx))
    ws2.cell(row=i, column=2, value=row["收入"]).number_format = '£#,##0.00'
    ws2.cell(row=i, column=3, value=int(row["订单数"]))
    ws2.cell(row=i, column=4, value=int(row["客户数"]))
    ws2.cell(row=i, column=5, value=row["客单价"]).number_format = '£#,##0.00'

style_data(ws2, 4, 3 + len(monthly), 5)
ws2.column_dimensions["A"].width = 12
ws2.column_dimensions["B"].width = 14
ws2.column_dimensions["C"].width = 10
ws2.column_dimensions["D"].width = 10
ws2.column_dimensions["E"].width = 14

# 月度收入柱状图
chart = BarChart()
chart.type = "col"
chart.title = "月度收入趋势"
chart.y_axis.title = "收入 (£)"
chart.x_axis.title = "月份"
chart.style = 10
data = Reference(ws2, min_col=2, min_row=3, max_row=2 + len(monthly))
cats = Reference(ws2, min_col=1, min_row=4, max_row=3 + len(monthly))
chart.add_data(data, titles_from_data=True)
chart.set_categories(cats)
chart.shape = 4
ws2.add_chart(chart, "G3")


# ══════════════════════════════════════════
# Sheet 3: 国家分析
# ══════════════════════════════════════════
ws3 = wb.create_sheet("国家分析")
ws3.sheet_properties.tabColor = BRAND

add_title(ws3, 1, 1, "国家/地区分析")
ws3.merge_cells("A1:F1")

country = df.groupby("Country").agg(
    收入=("Revenue", "sum"),
    订单数=("InvoiceNo", "nunique"),
    客户数=("CustomerID", "nunique"),
    销量=("Quantity", "sum"),
).sort_values("收入", ascending=False)
country["客单价"] = round(country["收入"] / country["订单数"], 2)
country["收入占比"] = round(country["收入"] / country["收入"].sum() * 100, 1)

headers3 = ["国家", "收入", "订单数", "客户数", "客单价", "收入占比%"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=3, column=i, value=h)
style_header(ws3, 3, 6)

for i, (idx, row) in enumerate(country.iterrows(), 4):
    ws3.cell(row=i, column=1, value=idx)
    ws3.cell(row=i, column=2, value=row["收入"]).number_format = '£#,##0.00'
    ws3.cell(row=i, column=3, value=int(row["订单数"]))
    ws3.cell(row=i, column=4, value=int(row["客户数"]))
    ws3.cell(row=i, column=5, value=row["客单价"]).number_format = '£#,##0.00'
    ws3.cell(row=i, column=6, value=row["收入占比"])

style_data(ws3, 4, 3 + len(country), 6)
ws3.column_dimensions["A"].width = 18
ws3.column_dimensions["B"].width = 14
ws3.column_dimensions["C"].width = 10
ws3.column_dimensions["D"].width = 10
ws3.column_dimensions["E"].width = 14
ws3.column_dimensions["F"].width = 12

# Top 10 国家收入柱状图
chart3 = BarChart()
chart3.type = "bar"
chart3.title = "Top 10 国家收入"
chart3.style = 10
data3 = Reference(ws3, min_col=2, min_row=3, max_row=13)
cats3 = Reference(ws3, min_col=1, min_row=4, max_row=13)
chart3.add_data(data3, titles_from_data=True)
chart3.set_categories(cats3)
ws3.add_chart(chart3, "H3")


# ══════════════════════════════════════════
# Sheet 4: 商品分析
# ══════════════════════════════════════════
ws4 = wb.create_sheet("商品分析")
ws4.sheet_properties.tabColor = BRAND

add_title(ws4, 1, 1, "商品分析 (Top 30)")
ws4.merge_cells("A1:E1")

products = df.groupby(["StockCode", "Description"]).agg(
    收入=("Revenue", "sum"),
    销量=("Quantity", "sum"),
    订单数=("InvoiceNo", "nunique"),
).sort_values("收入", ascending=False).head(30)
products["均价"] = round(products["收入"] / products["销量"], 2)
products = products.round(2)

headers4 = ["商品编码", "描述", "收入", "销量", "订单数", "均价"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=3, column=i, value=h)
style_header(ws4, 3, 6)

for i, (idx, row) in enumerate(products.iterrows(), 4):
    ws4.cell(row=i, column=1, value=idx[0])
    ws4.cell(row=i, column=2, value=idx[1])
    ws4.cell(row=i, column=3, value=row["收入"]).number_format = '£#,##0.00'
    ws4.cell(row=i, column=4, value=int(row["销量"]))
    ws4.cell(row=i, column=5, value=int(row["订单数"]))
    ws4.cell(row=i, column=6, value=row["均价"]).number_format = '£#,##0.00'

style_data(ws4, 4, 3 + len(products), 6)
ws4.column_dimensions["A"].width = 12
ws4.column_dimensions["B"].width = 35
ws4.column_dimensions["C"].width = 14
ws4.column_dimensions["D"].width = 10
ws4.column_dimensions["E"].width = 10
ws4.column_dimensions["F"].width = 10


# ══════════════════════════════════════════
# Sheet 5: 客户分层
# ══════════════════════════════════════════
ws5 = wb.create_sheet("客户分层")
ws5.sheet_properties.tabColor = BRAND

add_title(ws5, 1, 1, "客户分层 (RFM)")
ws5.merge_cells("A1:E1")

ref_date = df["InvoiceDate"].max() + pd.Timedelta(days=1)
rfm = df.groupby("CustomerID").agg(
    Recency=("InvoiceDate", lambda x: (ref_date - x.max()).days),
    Frequency=("InvoiceNo", "nunique"),
    Monetary=("Revenue", "sum"),
)


def rfm_segment(row):
    if row["Recency"] <= 30 and row["Frequency"] >= 10:
        return "高价值活跃"
    elif row["Recency"] <= 90 and row["Frequency"] >= 5:
        return "中等活跃"
    elif row["Recency"] <= 180:
        return "低频但近"
    elif row["Frequency"] >= 5:
        return "曾活跃但流失"
    else:
        return "低频/沉睡"


rfm["Segment"] = rfm.apply(rfm_segment, axis=1)
segments = rfm.groupby("Segment").agg(
    客户数=("Monetary", "count"),
    人均消费=("Monetary", "mean"),
    总收入=("Monetary", "sum"),
).sort_values("总收入", ascending=False)
segments["客户占比%"] = round(segments["客户数"] / segments["客户数"].sum() * 100, 1)
segments["收入占比%"] = round(segments["总收入"] / segments["总收入"].sum() * 100, 1)
segments = segments.round(2)

headers5 = ["分层", "客户数", "客户占比%", "人均消费", "总收入", "收入占比%"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, 6)

for i, (idx, row) in enumerate(segments.iterrows(), 4):
    ws5.cell(row=i, column=1, value=idx)
    ws5.cell(row=i, column=2, value=int(row["客户数"]))
    ws5.cell(row=i, column=3, value=row["客户占比%"])
    ws5.cell(row=i, column=4, value=row["人均消费"]).number_format = '£#,##0.00'
    ws5.cell(row=i, column=5, value=row["总收入"]).number_format = '£#,##0.00'
    ws5.cell(row=i, column=6, value=row["收入占比%"])

style_data(ws5, 4, 3 + len(segments), 6)
ws5.column_dimensions["A"].width = 14
ws5.column_dimensions["B"].width = 10
ws5.column_dimensions["C"].width = 12
ws5.column_dimensions["D"].width = 14
ws5.column_dimensions["E"].width = 14
ws5.column_dimensions["F"].width = 12

# 收入占比柱状图
chart5 = BarChart()
chart5.type = "bar"
chart5.title = "各分层收入占比"
chart5.style = 10
data5 = Reference(ws5, min_col=6, min_row=3, max_row=3 + len(segments))
cats5 = Reference(ws5, min_col=1, min_row=4, max_row=3 + len(segments))
chart5.add_data(data5, titles_from_data=True)
chart5.set_categories(cats5)
ws5.add_chart(chart5, "A10")


# ══════════════════════════════════════════
# Sheet 6: 分析看板 (Dashboard)
# ══════════════════════════════════════════
ws6 = wb.create_sheet("分析看板")
ws6.sheet_properties.tabColor = BRAND

# KPI 行
dashboard_title = ws6.cell(row=1, column=1, value="📊 在线零售数据分析看板")
dashboard_title.font = Font(name="Noto Sans SC", bold=True, color=BRAND, size=16)
ws6.merge_cells("A1:H1")

kpis = [
    ("总收入", "£8,911,408", "12个月"),
    ("总客户数", "4,338", "有购买记录"),
    ("复购率", "65.6%", "多次购买占比"),
    ("客单价 AOV", "£480.87", "每单平均"),
]
for i, (label, val, note) in enumerate(kpis):
    c = i * 2 + 1
    cell_l = ws6.cell(row=3, column=c, value=label)
    cell_l.font = Font(name="Noto Sans SC", bold=True, color=STONE, size=10)
    cell_v = ws6.cell(row=4, column=c, value=val)
    cell_v.font = Font(name="JetBrains Mono", bold=True, color=BRAND, size=18)
    cell_n = ws6.cell(row=5, column=c, value=note)
    cell_n.font = Font(name="Noto Sans SC", color=STONE, size=9)
    ws6.merge_cells(start_row=3, start_column=c, end_row=3, end_column=c + 1)
    ws6.merge_cells(start_row=4, start_column=c, end_row=4, end_column=c + 1)
    ws6.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)
    for r in [3, 4, 5]:
        for cc in [c, c + 1]:
            ws6.cell(row=r, column=cc).fill = kpi_fill

# 核心发现
ws6.cell(row=7, column=1, value="核心发现").font = Font(name="Noto Sans SC", bold=True, color=BRAND, size=13)
ws6.merge_cells("A7:H7")

findings = [
    "1. 英国市场占 82% 收入，但荷兰客单价 £3,040 是英国 £439 的 7 倍",
    "2. 11 月收入 £1.16M 是 2 月 £0.45M 的 2.6 倍，圣诞季效应显著",
    "3. 高价值客户 346 人（8%）贡献了 47.8% 的收入，杠杆集中",
    "4. Top 10 商品合计不到 10% 收入，典型长尾零售模型",
    "5. 复购率 65.6%，但仍有 34.4% 的客户只买过一次",
]
for i, f in enumerate(findings, 8):
    ws6.cell(row=i, column=1, value=f).font = body_font
    ws6.merge_cells(start_row=i, start_column=1, end_row=i, end_column=8)

# 思维模型总结
ws6.cell(row=14, column=1, value="思维模型总结").font = Font(name="Noto Sans SC", bold=True, color=BRAND, size=13)
ws6.merge_cells("A14:H14")

models = [
    ("分层差异", "英国 vs 海外，高活跃 vs 沉睡客户——不要信平均值"),
    ("杠杆点识别", "8% 客户贡献 48% 收入，资源应集中"),
    ("分解思维", "年收入按月拆看季节波动，按国家拆看市场差异"),
]
hdr6 = ["思维模型", "核心洞察"]
for i, h in enumerate(hdr6, 1):
    cell = ws6.cell(row=15, column=i, value=h)
    cell.font = hdr_font
    cell.fill = hdr_fill
    cell.border = thin_border
ws6.merge_cells("B15:C15")

for i, (m, d) in enumerate(models, 16):
    ws6.cell(row=i, column=1, value=m).font = Font(name="Noto Sans SC", bold=True, size=10)
    ws6.cell(row=i, column=2, value=d).font = body_font
    ws6.merge_cells(start_row=i, start_column=2, end_row=i, end_column=3)
    for c in range(1, 4):
        ws6.cell(row=i, column=c).border = thin_border

# 行动建议
ws6.cell(row=20, column=1, value="行动建议").font = Font(name="Noto Sans SC", bold=True, color=BRAND, size=13)
ws6.merge_cells("A20:H20")

actions = [
    ("优先级", "行动", "预期效果"),
    ("P0", "高价值客户 VIP 专项维护", "稳住 48% 收入基本盘"),
    ("P0", "9-10 月提前布局圣诞季库存与营销", "抓住 2.6 倍旺季增量"),
    ("P1", "海外市场小规模测试（荷兰/爱尔兰）", "验证高客单价市场扩展"),
    ("P1", "单次客户 30 天内 follow-up", "提升 34.4% 单次客户的复购"),
    ("P2", "长尾商品策略：高频小物+高利润特色", "优化品类结构"),
]
for i, (p, act, effect) in enumerate(actions, 21):
    ws6.cell(row=i, column=1, value=p).font = Font(name="JetBrains Mono", bold=True, size=10, color=BRAND if p == "P0" else STONE)
    ws6.cell(row=i, column=2, value=act).font = body_font
    ws6.cell(row=i, column=3, value=effect).font = Font(name="Noto Sans SC", size=9, color=STONE)
    ws6.merge_cells(start_row=i, start_column=2, end_row=i, end_column=4)
    ws6.merge_cells(start_row=i, start_column=5, end_row=i, end_column=7)
    for c in range(1, 8):
        ws6.cell(row=i, column=c).border = thin_border

ws6.column_dimensions["A"].width = 10
ws6.column_dimensions["B"].width = 35
ws6.column_dimensions["C"].width = 20
ws6.column_dimensions["D"].width = 10
for c in "EFGH":
    ws6.column_dimensions[c].width = 12

# ── 保存 ──
out_path = os.path.join(OUT_DIR, "online_retail_analysis.xlsx")
wb.save(out_path)
print(f"✅ Excel saved: {out_path}")
