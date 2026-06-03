"""
Marketing Campaign Analysis — Excel Deliverable
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from copy import copy

# ── Data ──
df = pd.read_csv("/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/work/kaggle_datasets/downloads/marketing_campaign_performance/KAG_conversion_data.csv")

df["CPC"] = df["Spent"] / df["Clicks"].replace(0, np.nan)
df["CPM"] = df["Spent"] / df["Impressions"].replace(0, np.nan) * 1000
df["CTR"] = df["Clicks"] / df["Impressions"].replace(0, np.nan) * 100
df["Conversion_Rate"] = df["Total_Conversion"] / df["Clicks"].replace(0, np.nan) * 100

wb = Workbook()

# ── Styles ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_font = Font(bold=True, size=14, color="1F3864")
kpi_font = Font(bold=True, size=20, color="2F5496")
kpi_label_font = Font(size=10, color="666666")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_data(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

def add_chart(ws, data_ref, cat_ref, row, col, title, chart_type="bar"):
    if chart_type == "bar":
        chart = BarChart()
        chart.type = "col"
    else:
        chart = PieChart()
    chart.title = title
    chart.style = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4
    ws.add_chart(chart, f"{chr(64 + col)}{row}")

# ════════════════════════════════════════
# Sheet 1: 整体概览
# ════════════════════════════════════════
ws1 = wb.active
ws1.title = "整体概览"
ws1.merge_cells("A1:F1")
ws1["A1"] = "Marketing Campaign Performance — 关键指标总览"
ws1["A1"].font = title_font

overview = pd.DataFrame([
    ("总花费 (Spent)", f"${df['Spent'].sum():,.2f}", "所有广告花费总和"),
    ("总展示 (Impressions)", f"{df['Impressions'].sum():,}", "广告被展示总次数"),
    ("总点击 (Clicks)", f"{df['Clicks'].sum():,}", "广告被点击总次数"),
    ("总转化 (Conversions)", f"{df['Total_Conversion'].sum():,}", "总转化数"),
    ("审核通过转化", f"{df['Approved_Conversion'].sum():,}", "审核通过的转化数"),
    ("整体 CPC", f"${df['Spent'].sum()/df['Clicks'].sum():.4f}", "平均每次点击成本"),
    ("整体 CPM", f"${df['Spent'].sum()/df['Impressions'].sum()*1000:.2f}", "每千次展示成本"),
    ("整体 CTR", f"{df['Clicks'].sum()/df['Impressions'].sum()*100:.2f}%", "点击率"),
    ("转化率", f"{df['Total_Conversion'].sum()/df['Clicks'].sum()*100:.2f}%", "点击→转化比例"),
    ("每转化成本", f"${df['Spent'].sum()/df['Total_Conversion'].sum():.2f}", "平均获取一个转化的成本"),
], columns=["指标", "数值", "说明"])

for r in dataframe_to_rows(overview, index=False, header=True):
    ws1.append(r)
style_header(ws1, 3, 3)
style_data(ws1, 4, 3 + len(overview), 3)

# ════════════════════════════════════════
# Sheet 2: 广告系列对比
# ════════════════════════════════════════
ws2 = wb.create_sheet("广告系列对比")
ws2.merge_cells("A1:I1")
ws2["A1"] = "广告系列 (Campaign) 对比分析"
ws2["A1"].font = title_font

campaigns = df.groupby("xyz_campaign_id").agg(
    广告数=("ad_id", "count"),
    花费=("Spent", "sum"),
    展示=("Impressions", "sum"),
    点击=("Clicks", "sum"),
    转化=("Total_Conversion", "sum"),
    审核通过=("Approved_Conversion", "sum"),
)
campaigns["CPC"] = campaigns["花费"] / campaigns["点击"]
campaigns["CTR"] = campaigns["点击"] / campaigns["展示"] * 100
campaigns["转化率"] = campaigns["转化"] / campaigns["点击"] * 100
campaigns["每转化成本"] = campaigns["花费"] / campaigns["转化"]
campaigns = campaigns.reset_index()

for r in dataframe_to_rows(campaigns, index=False, header=True):
    ws2.append(r)
style_header(ws2, 3, len(campaigns.columns))
style_data(ws2, 4, 3 + len(campaigns), len(campaigns.columns))

add_chart(ws2,
    Reference(ws2, min_col=3, min_row=3, max_col=3, max_row=3 + len(campaigns)),
    Reference(ws2, min_col=1, min_row=4, max_row=3 + len(campaigns)),
    row=8, col=1, title="各广告系列花费对比"
)
add_chart(ws2,
    Reference(ws2, min_col=8, min_row=3, max_col=8, max_row=3 + len(campaigns)),
    Reference(ws2, min_col=1, min_row=4, max_row=3 + len(campaigns)),
    row=8, col=5, title="各广告系列每转化成本对比"
)

# ════════════════════════════════════════
# Sheet 3: 年龄组分析
# ════════════════════════════════════════
ws3 = wb.create_sheet("年龄组分析")
ws3.merge_cells("A1:H1")
ws3["A1"] = "年龄组表现分析"
ws3["A1"].font = title_font

age = df.groupby("age").agg(
    花费=("Spent", "sum"),
    展示=("Impressions", "sum"),
    点击=("Clicks", "sum"),
    转化=("Total_Conversion", "sum"),
    审核通过=("Approved_Conversion", "sum"),
).round(2)
age["CPC"] = age["花费"] / age["点击"]
age["CTR"] = age["点击"] / age["展示"] * 100
age["每转化成本"] = age["花费"] / age["转化"]
age = age.reset_index()

for r in dataframe_to_rows(age, index=False, header=True):
    ws3.append(r)
style_header(ws3, 3, len(age.columns))
style_data(ws3, 4, 3 + len(age), len(age.columns))

add_chart(ws3,
    Reference(ws3, min_col=2, min_row=3, max_col=2, max_row=3 + len(age)),
    Reference(ws3, min_col=1, min_row=4, max_row=3 + len(age)),
    row=9, col=1, title="各年龄组花费"
)
add_chart(ws3,
    Reference(ws3, min_col=8, min_row=3, max_col=8, max_row=3 + len(age)),
    Reference(ws3, min_col=1, min_row=4, max_row=3 + len(age)),
    row=9, col=4, title="各年龄组每转化成本"
)
add_chart(ws3,
    Reference(ws3, min_col=7, min_row=3, max_col=7, max_row=3 + len(age)),
    Reference(ws3, min_col=1, min_row=4, max_row=3 + len(age)),
    row=9, col=7, title="各年龄组 CTR"
)

# ════════════════════════════════════════
# Sheet 4: 性别分析
# ════════════════════════════════════════
ws4 = wb.create_sheet("性别分析")
ws4.merge_cells("A1:H1")
ws4["A1"] = "性别表现分析"
ws4["A1"].font = title_font

gender = df.groupby("gender").agg(
    花费=("Spent", "sum"),
    展示=("Impressions", "sum"),
    点击=("Clicks", "sum"),
    转化=("Total_Conversion", "sum"),
    审核通过=("Approved_Conversion", "sum"),
).round(2)
gender["CPC"] = gender["花费"] / gender["点击"]
gender["CTR"] = gender["点击"] / gender["展示"] * 100
gender["每转化成本"] = gender["花费"] / gender["转化"]
gender = gender.reset_index()

for r in dataframe_to_rows(gender, index=False, header=True):
    ws4.append(r)
style_header(ws4, 3, len(gender.columns))
style_data(ws4, 4, 3 + len(gender), len(gender.columns))

# ════════════════════════════════════════
# Sheet 5: 年龄×性别交叉分析
# ════════════════════════════════════════
ws5 = wb.create_sheet("年龄×性别交叉")
ws5.merge_cells("A1:G1")
ws5["A1"] = "年龄×性别 交叉分析"
ws5["A1"].font = title_font

cross = df.groupby(["age", "gender"]).agg(
    花费=("Spent", "sum"),
    展示=("Impressions", "sum"),
    点击=("Clicks", "sum"),
    转化=("Total_Conversion", "sum"),
).round(2)
cross["CPC"] = cross["花费"] / cross["点击"]
cross["CTR"] = cross["点击"] / cross["展示"] * 100
cross["每转化成本"] = cross["花费"] / cross["转化"]
cross = cross.reset_index()

for r in dataframe_to_rows(cross, index=False, header=True):
    ws5.append(r)
style_header(ws5, 3, len(cross.columns))
style_data(ws5, 4, 3 + len(cross), len(cross.columns))

# ════════════════════════════════════════
# Sheet 6: 兴趣分类 Top
# ════════════════════════════════════════
ws6 = wb.create_sheet("兴趣分类分析")
ws6.merge_cells("A1:F1")
ws6["A1"] = "兴趣分类 Top 15 (按花费)"
ws6["A1"].font = title_font

interest = df.groupby("interest").agg(
    花费=("Spent", "sum"),
    展示=("Impressions", "sum"),
    点击=("Clicks", "sum"),
    转化=("Total_Conversion", "sum"),
).sort_values("花费", ascending=False)
interest["CPC"] = interest["花费"] / interest["点击"]
interest["每转化成本"] = interest["花费"] / interest["转化"]
interest = interest.head(15).reset_index()

for r in dataframe_to_rows(interest, index=False, header=True):
    ws6.append(r)
style_header(ws6, 3, len(interest.columns))
style_data(ws6, 4, 3 + len(interest), len(interest.columns))

# ════════════════════════════════════════
# Sheet 7: 分析看板
# ════════════════════════════════════════
ws7 = wb.create_sheet("分析看板")
ws7.merge_cells("A1:F1")
ws7["A1"] = "Marketing Campaign 分析看板"
ws7["A1"].font = Font(bold=True, size=16, color="1F3864")

# KPI row
kpis = [
    ("总花费", f"${df['Spent'].sum():,.0f}"),
    ("总展示", f"{df['Impressions'].sum()/1e6:.1f}M"),
    ("总点击", f"{df['Clicks'].sum():,}"),
    ("总转化", f"{df['Total_Conversion'].sum():,}"),
    ("每转化成本", f"${df['Spent'].sum()/df['Total_Conversion'].sum():.2f}"),
    ("CTR", f"{df['Clicks'].sum()/df['Impressions'].sum()*100:.2f}%"),
]
for i, (label, val) in enumerate(kpis):
    col = i * 2 + 1
    ws7.cell(row=3, column=col, value=label).font = kpi_label_font
    ws7.cell(row=4, column=col, value=val).font = kpi_font
    ws7.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col + 1)
    ws7.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col + 1)

# Core findings
ws7.cell(row=6, column=1, value="核心发现").font = Font(bold=True, size=12, color="2F5496")
findings = [
    f"广告系列 1178 占总花费的 94.8% (${df[df['xyz_campaign_id']==1178]['Spent'].sum():,.0f})，但其转化率仅 7.4%，远低于系列 916 的 51.3%",
    f"年龄组 45-49 花费最高 (${df[df['age']=='45-49']['Spent'].sum():,.0f})，但每转化成本也最高 ($30.34)",
    f"男性转化成本 ($14.94) 低于女性 ($21.00)，但女性 CTR (2.08%) 高于男性 (1.45%)",
    f"兴趣分类 16 花费最高 ($8,085)，但兴趣分类 15 每转化成本最低 ($13.32)",
]
for j, f in enumerate(findings):
    ws7.cell(row=7 + j, column=1, value=f"• {f}")

# Recommendations
ws7.cell(row=13, column=1, value="行动建议").font = Font(bold=True, size=12, color="2F5496")
recs = [
    "重点优化广告系列 1178 的转化效率，现有 7.4% 转化率远低于其他系列",
    "女性 30-34 组转化最多 (619)，可加大该组合预算，同时优化定向精准度",
    "兴趣分类 15、29、10 的每转化成本低于平均水平，建议增加预算分配",
    "年龄 45-49 组花费最高但效率最低，需检查素材和出价策略是否匹配",
]
for k, r in enumerate(recs):
    ws7.cell(row=14 + k, column=1, value=f"→ {r}")

# Thinking models
ws7.cell(row=19, column=1, value="思维模型").font = Font(bold=True, size=12, color="2F5496")
models = [
    "1. 分解思维 — 总花费 $58,705 按广告系列、年龄、性别、兴趣拆解，发现系列 1178 占 94.8% 但效率最低",
    "2. 分层差异 — 不同年龄组的 CTR (1.4%~2.2%) 和每转化成本 ($10.66~$30.34) 差异达 3 倍",
    "3. 杠杆点识别 — 女性 30-34 贡献最多转化(619)且每转化成本中等($12.30)，是优先优化组合",
    "4. 间接推断 — 审核通过率(33.1%)可作为转化质量的代理指标",
    "5. 约束 vs 偏好 — 年龄组间的转化成本差异可能反映产品-市场匹配度，而非投放策略问题",
]
for m in models:
    ws7.cell(row=20 + models.index(m), column=1, value=m)

# Column widths
from openpyxl.utils import get_column_letter
for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
    for col_idx in range(1, ws.max_column + 1):
        col_vals = []
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                col_vals.append(len(str(v)))
        max_len = max(col_vals, default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

# ── Save ──
out_path = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output/marketing_analysis.xlsx"
wb.save(out_path)
print(f"Excel saved: {out_path}")
