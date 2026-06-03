"""
SaaS Churn Analysis — Excel Deliverable
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

BASE = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/work/kaggle_datasets/downloads/saas_subscription_churn"

accounts = pd.read_csv(f"{BASE}/ravenstack_accounts.csv")
subscriptions = pd.read_csv(f"{BASE}/ravenstack_subscriptions.csv")
churn_events = pd.read_csv(f"{BASE}/ravenstack_churn_events.csv")

wb = Workbook()
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
title_font = Font(bold=True, size=14, color="1F3864")
kpi_font = Font(bold=True, size=18, color="2F5496")
kpi_label_font = Font(size=9, color="666666")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header(ws, row, max_col):
    for c in range(1, max_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

def style_data(ws, start_row, end_row, max_col):
    for r in range(start_row, end_row + 1):
        for c in range(1, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

def add_chart(ws, data_ref, cat_ref, row, col, title):
    chart = BarChart()
    chart.type = "col"
    chart.title = title
    chart.style = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    chart.shape = 4
    ws.add_chart(chart, f"{chr(64 + col)}{row}")

total = len(accounts)
churned = accounts["churn_flag"].sum()
churn_rate = churned / total * 100
total_arr = subscriptions["arr_amount"].sum()
churned_accounts = accounts[accounts["churn_flag"] == True]["account_id"]
churned_arr = subscriptions[subscriptions["account_id"].isin(churned_accounts)]["arr_amount"].sum()

# Sheet 1: Overall
ws1 = wb.active
ws1.title = "整体概览"
ws1.merge_cells("A1:C1")
ws1["A1"] = "SaaS Churn Analysis — 关键指标"
ws1["A1"].font = title_font

kpis = [
    ("总客户数", f"{total}", "Accounts"),
    ("流失客户", f"{churned} ({churn_rate:.1f}%)", "Churned"),
    ("留存客户", f"{total-churned} ({100-churn_rate:.1f}%)", "Retained"),
    ("总 ARR", f"${total_arr:,.0f}", "Annual Recurring Revenue"),
    ("流失 ARR", f"${churned_arr:,.0f} ({churned_arr/total_arr*100:.1f}%)", "Churned ARR"),
    ("平均 MRR", f"${subscriptions['mrr_amount'].mean():.0f}", "Avg Monthly per subscription"),
]
for i, (label, val, note) in enumerate(kpis):
    r = 3 + i
    ws1.cell(row=r, column=1, value=label).font = Font(bold=True)
    ws1.cell(row=r, column=2, value=val).font = kpi_font
    ws1.cell(row=r, column=3, value=note).font = kpi_label_font

# Sheet 2: Industry churn
ws2 = wb.create_sheet("行业流失率")
ws2.merge_cells("A1:D1")
ws2["A1"] = "各行业流失率"
ws2["A1"].font = title_font

ind = accounts.groupby("industry").agg(
    客户数=("account_id", "count"),
    流失数=("churn_flag", "sum"),
)
ind["流失率"] = ind["流失数"] / ind["客户数"] * 100
ind = ind.reset_index()
for r in dataframe_to_rows(ind, index=False, header=True):
    ws2.append(r)
style_header(ws2, 3, len(ind.columns))
style_data(ws2, 4, 3 + len(ind), len(ind.columns))
add_chart(ws2,
    Reference(ws2, min_col=4, min_row=3, max_col=4, max_row=3+len(ind)),
    Reference(ws2, min_col=1, min_row=4, max_row=3+len(ind)),
    8, 1, "各行业流失率")

# Sheet 3: Plan tier
ws3 = wb.create_sheet("套餐分析")
ws3.merge_cells("A1:E1")
ws3["A1"] = "套餐层级表现"
ws3["A1"].font = title_font

plan = accounts.groupby("plan_tier").agg(
    客户数=("account_id", "count"),
    流失数=("churn_flag", "sum"),
)
plan["流失率"] = plan["流失数"] / plan["客户数"] * 100
plan_rev = subscriptions.groupby("plan_tier").agg(
    订阅数=("subscription_id", "count"),
    MRR=("mrr_amount", "sum"),
    ARR=("arr_amount", "sum"),
).round(2)
plan = plan.join(plan_rev).reset_index()
for r in dataframe_to_rows(plan, index=False, header=True):
    ws3.append(r)
style_header(ws3, 3, len(plan.columns))
style_data(ws3, 4, 3 + len(plan), len(plan.columns))

# Sheet 4: Referral source
ws4 = wb.create_sheet("来源渠道")
ws4.merge_cells("A1:D1")
ws4["A1"] = "注册来源渠道分析"
ws4["A1"].font = title_font

ref = accounts.groupby("referral_source").agg(
    客户数=("account_id", "count"),
    流失数=("churn_flag", "sum"),
)
ref["流失率"] = ref["流失数"] / ref["客户数"] * 100
ref = ref.reset_index()
for r in dataframe_to_rows(ref, index=False, header=True):
    ws4.append(r)
style_header(ws4, 3, len(ref.columns))
style_data(ws4, 4, 3 + len(ref), len(ref.columns))
add_chart(ws4,
    Reference(ws4, min_col=4, min_row=3, max_col=4, max_row=3+len(ref)),
    Reference(ws4, min_col=1, min_row=4, max_row=3+len(ref)),
    8, 1, "各渠道来源流失率")

# Sheet 5: Churn reasons
ws5 = wb.create_sheet("流失原因")
ws5.merge_cells("A1:C1")
ws5["A1"] = "流失原因分布"
ws5["A1"].font = title_font

reasons = churn_events["reason_code"].value_counts().reset_index()
reasons.columns = ["原因", "次数"]
reasons["占比"] = reasons["次数"] / len(churn_events) * 100
for r in dataframe_to_rows(reasons, index=False, header=True):
    ws5.append(r)
style_header(ws5, 3, len(reasons.columns))
style_data(ws5, 4, 3 + len(reasons), len(reasons.columns))

# Sheet 6: Country
ws6 = wb.create_sheet("国家分布")
ws6.merge_cells("A1:D1")
ws6["A1"] = "各国流失率"
ws6["A1"].font = title_font

ctry = accounts.groupby("country").agg(
    客户数=("account_id", "count"),
    流失数=("churn_flag", "sum"),
)
ctry["流失率"] = ctry["流失数"] / ctry["客户数"] * 100
ctry = ctry.reset_index().sort_values("客户数", ascending=False)
for r in dataframe_to_rows(ctry, index=False, header=True):
    ws6.append(r)
style_header(ws6, 3, len(ctry.columns))
style_data(ws6, 4, 3 + len(ctry), len(ctry.columns))

# Sheet 7: Dashboard
ws7 = wb.create_sheet("分析看板")
ws7.merge_cells("A1:F1")
ws7["A1"] = "SaaS Churn 分析看板"
ws7["A1"].font = Font(bold=True, size=16, color="1F3864")

ws7.cell(row=3, column=1, value="核心发现").font = Font(bold=True, size=12, color="2F5496")
findings = [
    f"整体流失率: {churn_rate:.1f}% — 500 个客户中 110 个流失，流失 ARR 占比 {churned_arr/total_arr*100:.1f}%",
    f"DevTools 行业流失率最高 (31.0%)，远高于 EdTech (16.5%) 和 Cybersecurity (16.0%)",
    f"活动来源 (event) 渠道客户流失率 30.2%，合作伙伴 (partner) 渠道仅 14.6%",
    f"主要流失原因: features (19.0%)、support (17.3%)、budget (17.3%) — 产品和价格是核心",
    f"流失客户工单升级率 (0.06) 高于留存客户 (0.04) — 工单升级是潜在提前预警信号",
]
for j, f in enumerate(findings):
    ws7.cell(row=4 + j, column=1, value=f"• {f}")

ws7.cell(row=10, column=1, value="行动建议").font = Font(bold=True, size=12, color="2F5496")
recs = [
    "优先改善 DevTools 行业的产品体验与支持响应 — 该行业占总客户 22.6% 但流失占 31.8%",
    "优化活动获客后的激活流程 — event 渠道客户流失率 30.2% 说明留存后劲不足",
    "展开流失客户调研，重点分析 features 和 support 两类原因的深层根因",
    "利用工单升级作为提前预警信号 — escalation_rate 从 0.04 升高到 0.06 即触发客户成功干预",
    "试用客户转化流程需优化 — 试用流失率 25.8% vs 非试用 21.1%",
]
for k, r2 in enumerate(recs):
    ws7.cell(row=11 + k, column=1, value=f"→ {r2}")

for ws in [ws1, ws2, ws3, ws4, ws5, ws6, ws7]:
    for col_idx in range(1, ws.max_column + 1):
        col_vals = []
        for row_idx in range(1, ws.max_row + 1):
            v = ws.cell(row=row_idx, column=col_idx).value
            if v is not None:
                col_vals.append(len(str(v)))
        max_len = max(col_vals, default=10)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 50)

out = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output/saas_churn_analysis.xlsx"
wb.save(out)
print(f"Excel saved: {out}")
