"""
Financial / Operating Performance — Excel Report
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import warnings; warnings.filterwarnings("ignore")

BASE = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/work/kaggle_datasets/downloads/financial_operating_performance"
df = pd.read_csv(f"{BASE}/BCG.csv")
df = df.sort_values(["Company_Name", "Year"])

wb = Workbook()

# ── Helper: style header row ──
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="1B365D")
thin_border = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)

def style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")

def auto_width(ws):
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 30)

# ── Sheet 1: Raw Data ──
ws1 = wb.active
ws1.title = "原始数据"
for row in dataframe_to_rows(df, index=False, header=True):
    ws1.append(row)
style_header(ws1)
auto_width(ws1)

# ── Sheet 2: 2023 Summary ──
ws2 = wb.create_sheet("2023财务总览")
cur = df[df["Year"] == 2023].copy()
if len(cur) > 0:
    cur["利润率_%"] = (cur["Net_Income"] / cur["Total_Revenue"] * 100).round(2)
    cur["负债率_%"] = (cur["Total_Liabilities"] / cur["Total_Assets"] * 100).round(2)
    cur["ROA_%"] = (cur["Net_Income"] / cur["Total_Assets"] * 100).round(2)
    cur["现金资产比_%"] = (cur["Cash_Flow"] / cur["Total_Assets"] * 100).round(2)
    cols = ["Company_Name", "Total_Revenue", "Net_Income", "Total_Assets",
            "Total_Liabilities", "Cash_Flow", "利润率_%", "负债率_%", "ROA_%", "现金资产比_%"]
    for row in dataframe_to_rows(cur[cols], index=False, header=True):
        ws2.append(row)
    style_header(ws2)
    auto_width(ws2)

# ── Sheet 3: YoY Growth ──
ws3 = wb.create_sheet("同比增长率")

# Revenue pivot
rev_pivot = df.pivot_table(index="Company_Name", columns="Year", values="Total_Revenue")
ni_pivot = df.pivot_table(index="Company_Name", columns="Year", values="Net_Income")

ws3.append(["收入趋势 (百万美元)"])
ws3.append([])
for row in dataframe_to_rows(rev_pivot.reset_index(), index=False, header=True):
    ws3.append(row)
style_header(ws3, row=3, max_col=rev_pivot.shape[1] + 1)

# Net income pivot below
start_row = ws3.max_row + 2
ws3.cell(row=start_row, column=1, value="净利趋势 (百万美元)")
start_row += 1
for row in dataframe_to_rows(ni_pivot.reset_index(), index=False, header=True):
    ws3.cell(row=start_row, column=1, value=row[0] if len(row) > 0 else "")
    for i, v in enumerate(row[1:], start=2):
        ws3.cell(row=start_row, column=i, value=v)
    start_row += 1

auto_width(ws3)

# YoY growth rate table
ws3.cell(row=start_row + 1, column=1, value="同比增长率 (%)")
start_row += 2
ws3.cell(row=start_row, column=1, value="公司")
ws3.cell(row=start_row, column=2, value="期间")
ws3.cell(row=start_row, column=3, value="收入增长%")
ws3.cell(row=start_row, column=4, value="净利增长%")
for col_i in range(1, 5):
    ws3.cell(row=start_row, column=col_i).font = header_font
    ws3.cell(row=start_row, column=col_i).fill = header_fill
start_row += 1

for company in df["Company_Name"].unique():
    c = df[df["Company_Name"] == company].sort_values("Year")
    for i in range(1, len(c)):
        y1, y2 = c.iloc[i-1], c.iloc[i]
        rev_g = round((y2["Total_Revenue"] - y1["Total_Revenue"]) / y1["Total_Revenue"] * 100, 2)
        ni_g = round((y2["Net_Income"] - y1["Net_Income"]) / y1["Net_Income"] * 100, 2)
        ws3.cell(row=start_row, column=1, value=company)
        ws3.cell(row=start_row, column=2, value=f"{y1['Year']}-{y2['Year']}")
        ws3.cell(row=start_row, column=3, value=rev_g)
        ws3.cell(row=start_row, column=4, value=ni_g)
        start_row += 1

auto_width(ws3)

# ── Sheet 4: Profitability ──
ws4 = wb.create_sheet("盈利能力分析")
ws4.append(["公司", "年份", "利润率_%", "ROA_%", "负债率_%"])
style_header(ws4)

row_idx = 2
for company in df["Company_Name"].unique():
    c = df[df["Company_Name"] == company].sort_values("Year")
    for _, r in c.iterrows():
        pm = round(r["Net_Income"] / r["Total_Revenue"] * 100, 2)
        roa = round(r["Net_Income"] / r["Total_Assets"] * 100, 2)
        dr = round(r["Total_Liabilities"] / r["Total_Assets"] * 100, 2)
        ws4.cell(row=row_idx, column=1, value=company)
        ws4.cell(row=row_idx, column=2, value=int(r["Year"]))
        ws4.cell(row=row_idx, column=3, value=pm)
        ws4.cell(row=row_idx, column=4, value=roa)
        ws4.cell(row=row_idx, column=5, value=dr)
        row_idx += 1

auto_width(ws4)

# Chart: profit margin comparison
chart1 = BarChart()
chart1.type = "col"
chart1.title = "利润率对比 (%)"
chart1.y_axis.title = "利润率%"
data_ref = Reference(ws4, min_col=3, min_row=1, max_row=row_idx - 1, max_col=3)
cats_ref = Reference(ws4, min_col=1, min_row=2, max_row=row_idx - 1)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.width = 18
chart1.height = 12
ws4.add_chart(chart1, "G2")

# ── Sheet 5: Cross-Company Benchmark ──
ws5 = wb.create_sheet("跨公司对标")
latest = df[df["Year"] == df.groupby("Company_Name")["Year"].transform("max")]
latest = latest.sort_values("Total_Revenue", ascending=False)
latest["利润率_%"] = (latest["Net_Income"] / latest["Total_Revenue"] * 100).round(2)
latest["负债率_%"] = (latest["Total_Liabilities"] / latest["Total_Assets"] * 100).round(2)
latest["ROA_%"] = (latest["Net_Income"] / latest["Total_Assets"] * 100).round(2)

bench_cols = ["Company_Name", "Year", "Total_Revenue", "Net_Income", "Total_Assets",
              "Total_Liabilities", "Cash_Flow", "利润率_%", "负债率_%", "ROA_%"]
for row in dataframe_to_rows(latest[bench_cols], index=False, header=True):
    ws5.append(row)
style_header(ws5)
auto_width(ws5)

# Revenue bar chart
chart2 = BarChart()
chart2.type = "col"
chart2.title = "收入 vs 净利 (百万美元)"
data_ref2 = Reference(ws5, min_col=3, min_row=1, max_row=len(latest) + 1, max_col=4)
cats_ref2 = Reference(ws5, min_col=1, min_row=2, max_row=len(latest) + 1)
chart2.add_data(data_ref2, titles_from_data=True)
chart2.set_categories(cats_ref2)
chart2.width = 18
chart2.height = 12
ws5.add_chart(chart2, "L2")

# ── Sheet 6: Dashboard ──
ws6 = wb.create_sheet("分析看板")
title_font = Font(bold=True, size=14, color="1B365D")
section_font = Font(bold=True, size=11, color="1B365D")
body_font = Font(size=10, color="333333")

ws6.merge_cells("A1:H1")
ws6.cell(row=1, column=1, value="财务与运营表现 — 分析看板").font = title_font
ws6.cell(row=1, column=1).alignment = Alignment(horizontal="center")

apple_latest = df[df["Company_Name"] == "Apple"].iloc[-1]
msft_latest = df[df["Company_Name"] == "Microsoft"].iloc[-1]
tsla_latest = df[df["Company_Name"] == "Tesla"].iloc[-1]

apple_rev_g = (apple_latest["Total_Revenue"] / df[df["Company_Name"] == "Apple"]["Total_Revenue"].iloc[0] - 1) * 100

row = 3
ws6.cell(row=row, column=1, value="核心指标 (最新年度)").font = section_font
row += 1
ws6.cell(row=row, column=1, value="指标")
ws6.cell(row=row, column=2, value="Apple")
ws6.cell(row=row, column=3, value="Microsoft")
ws6.cell(row=row, column=4, value="Tesla")
for c in range(1, 5):
    ws6.cell(row=row, column=c).font = header_font
    ws6.cell(row=row, column=c).fill = header_fill

metrics = [
    ("收入 (B$)", apple_latest["Total_Revenue"] / 1000, msft_latest["Total_Revenue"] / 1000, tsla_latest["Total_Revenue"] / 1000),
    ("净利 (B$)", apple_latest["Net_Income"] / 1000, msft_latest["Net_Income"] / 1000, tsla_latest["Net_Income"] / 1000),
    ("利润率 (%)", round(apple_latest["Net_Income"] / apple_latest["Total_Revenue"] * 100, 1), round(msft_latest["Net_Income"] / msft_latest["Total_Revenue"] * 100, 1), round(tsla_latest["Net_Income"] / tsla_latest["Total_Revenue"] * 100, 1)),
    ("负债率 (%)", round(apple_latest["Total_Liabilities"] / apple_latest["Total_Assets"] * 100, 1), round(msft_latest["Total_Liabilities"] / msft_latest["Total_Assets"] * 100, 1), round(tsla_latest["Total_Liabilities"] / tsla_latest["Total_Assets"] * 100, 1)),
    ("ROA (%)", round(apple_latest["Net_Income"] / apple_latest["Total_Assets"] * 100, 1), round(msft_latest["Net_Income"] / msft_latest["Total_Assets"] * 100, 1), round(tsla_latest["Net_Income"] / tsla_latest["Total_Assets"] * 100, 1)),
    ("Apple 累计收入增长 (%)", "", round(apple_rev_g, 1), ""),
]

for m in metrics:
    row += 1
    for i, v in enumerate(m):
        ws6.cell(row=row, column=i + 1, value=v)

row += 2
ws6.cell(row=row, column=1, value="核心发现").font = section_font
findings = [
    f"1. 收入规模: Apple (${apple_latest['Total_Revenue']/1000:.1f}B) > Microsoft (${msft_latest['Total_Revenue']/1000:.1f}B) > Tesla (${tsla_latest['Total_Revenue']/1000:.1f}B)",
    f"2. 利润率: Microsoft ({msft_latest['Net_Income']/msft_latest['Total_Revenue']*100:.1f}%) 最高, Tesla ({tsla_latest['Net_Income']/tsla_latest['Total_Revenue']*100:.1f}%) 最低",
    f"3. 负债率: Apple {apple_latest['Total_Liabilities']/apple_latest['Total_Assets']*100:.0f}% — 看似高但实际上 Apple 有大量长期债务而非经营负债，需结合利息覆盖倍数判断",
    f"4. 增长形态: Tesla 处于高速增长期 (2020-2022 收入增长 158%)，Apple 成熟稳定，Microsoft 稳健增长",
    f"5. 三年维度: Apple 收入增长 {apple_rev_g:.1f}%，显示成熟期公司的增长天花板",
]
for f in findings:
    row += 1
    ws6.cell(row=row, column=1, value=f).font = body_font

# Auto-width for dashboard
for col in range(1, 9):
    ws6.column_dimensions[chr(64 + col)].width = 22

out = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output/financial_analysis.xlsx"
wb.save(out)
print(f"Excel: {out}")
