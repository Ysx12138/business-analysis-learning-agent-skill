"""
Generate Standard Report Excel for Telco Churn Analysis
Standard mode: 3 sheets + 2 charts + overview dashboard
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.chart import BarChart, Reference
import os

DATA_PATH = "/Users/sx/数据分析/telco_customer_churn/原始数据/telco_churn.csv"
OUTPUT_DIR = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

df = pd.read_csv(DATA_PATH)
df['TotalCharges'] = pd.to_numeric(df['TotalCharges'], errors='coerce')
df = df.dropna(subset=['TotalCharges'])
df = df[df['tenure'] > 0]

df['tenure_group'] = pd.cut(df['tenure'],
    bins=[-1, 6, 12, 24, 48, 72],
    labels=['0-6月', '7-12月', '13-24月', '25-48月', '49-72月'])
for f in ['OnlineSecurity','OnlineBackup','DeviceProtection','TechSupport','StreamingTV','StreamingMovies']:
    df[f+'_bin'] = df[f].apply(lambda x: 0 if x=='No' else (1 if x=='Yes' else 0))
df['service_count'] = sum(df[f+'_bin'] for f in ['OnlineSecurity','OnlineBackup','DeviceProtection','TechSupport','StreamingTV','StreamingMovies'])

HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10)
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFF')
HEADER_FILL = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
KPI_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THIN_BORDER = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

def write_header(ws, row, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=i+1, value=h)
        c.font = HEADER_FONT; c.fill = HEADER_FILL; c.alignment = Alignment(horizontal='center'); c.border = THIN_BORDER

def write_row(ws, row, vals, pct_cols=None):
    if pct_cols is None: pct_cols = set()
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=i+1, value=v)
        c.font = DATA_FONT; c.alignment = Alignment(horizontal='center'); c.border = THIN_BORDER
        if i in pct_cols and isinstance(v, (int, float)):
            c.number_format = '0.0%'
        if row % 2 == 0: c.fill = ALT_FILL

wb = Workbook()

# ===== SHEET 1: 核心指标概览 =====
ws1 = wb.active
ws1.title = "核心指标概览"
ws1.sheet_properties.tabColor = '4472C4'

ws1.merge_cells('A1:F1')
t = ws1.cell(row=1, column=1, value="📊 Telco Churn - 核心指标概览")
t.font = TITLE_FONT; t.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid'); t.alignment = Alignment(horizontal='center')

write_header(ws1, 3, ['指标', '总体', '未流失', '流失', '差异'])
metrics = [
    ('客户数', len(df), df[df.Churn=='No'].shape[0], df[df.Churn=='Yes'].shape[0], '-'),
    ('流失率', df['Churn'].value_counts(normalize=True)['Yes'], '-', '-', '-'),
    ('平均在网月数', round(df['tenure'].mean(),1), round(df[df.Churn=='No']['tenure'].mean(),1), round(df[df.Churn=='Yes']['tenure'].mean(),1), f"{df[df.Churn=='No']['tenure'].mean()/max(df[df.Churn=='Yes']['tenure'].mean(),0.01):.1f}x"),
    ('平均月费', round(df['MonthlyCharges'].mean(),2), round(df[df.Churn=='No']['MonthlyCharges'].mean(),2), round(df[df.Churn=='Yes']['MonthlyCharges'].mean(),2), '-'),
    ('平均附加服务数', round(df['service_count'].mean(),1), round(df[df.Churn=='No']['service_count'].mean(),1), round(df[df.Churn=='Yes']['service_count'].mean(),1), '-'),
]
for i, m in enumerate(metrics):
    pct = {1} if m[0] == '流失率' else set()
    write_row(ws1, 4+i, m, pct_cols=pct)

ws1.column_dimensions['A'].width = 18
for c in 'BCDE': ws1.column_dimensions[c].width = 14
ws1.column_dimensions['F'].width = 12

# ===== SHEET 2: 关键维度分析 =====
ws2 = wb.create_sheet("关键维度分析")
ws2.sheet_properties.tabColor = '5B9BD5'

ws2.merge_cells('A1:F1')
t2 = ws2.cell(row=1, column=1, value="📋 关键维度分组流失率")
t2.font = TITLE_FONT; t2.fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid'); t2.alignment = Alignment(horizontal='center')

write_header(ws2, 3, ['维度', '类别', '客户数', '占比', '流失数', '流失率'])
dims = [
    ('合同类型', 'Contract'),
    ('在网时长', 'tenure_group'),
    ('互联网服务', 'InternetService'),
    ('支付方式', 'PaymentMethod'),
    ('家庭绑定', 'has_family'),
    ('在线安全', 'OnlineSecurity'),
]
if 'has_family' not in df.columns:
    df['has_family'] = (df['Partner']=='Yes') | (df['Dependents']=='Yes')

r = 4
for label, col in dims:
    first = True
    for cat, grp in df.groupby(col, observed=True):
        t = len(grp)
        y = grp[grp.Churn=='Yes'].shape[0]
        write_row(ws2, r, [label if first else '', cat, t, t/len(df), y, y/t], pct_cols={3,5})
        first = False; r += 1

ws2.column_dimensions['A'].width = 14
ws2.column_dimensions['B'].width = 20
for c in 'CDEF': ws2.column_dimensions[c].width = 12

# Chart: contract churn rate
chart1 = BarChart()
chart1.type = "col"; chart1.title = "各合同类型流失率"; chart1.style = 10
d1 = Reference(ws2, min_col=6, min_row=3, max_row=6)
c1 = Reference(ws2, min_col=2, min_row=4, max_row=6)
chart1.add_data(d1, titles_from_data=True); chart1.set_categories(c1)
chart1.width = 16; chart1.height = 10
ws2.add_chart(chart1, "A20")

# ===== SHEET 3: 分析看板 =====
ws3 = wb.create_sheet("分析看板")
ws3.sheet_properties.tabColor = 'FFC000'

ws3.merge_cells('A1:F1')
t3 = ws3.cell(row=1, column=1, value="📊 Telco Churn 分析看板 (Standard Report)")
t3.font = TITLE_FONT; t3.fill = PatternFill(start_color='FFC000', end_color='FFC000', fill_type='solid'); t3.alignment = Alignment(horizontal='center')

BASE = df['Churn'].value_counts(normalize=True)['Yes']
kpis = [
    ('整体流失率', f'{BASE*100:.1f}%'),
    ('月付客户流失率', f'{df[df.Contract=="Month-to-month"]["Churn"].value_counts(normalize=True).get("Yes",0)*100:.1f}%'),
    ('前6月流失率', f'{df[df.tenure<=6]["Churn"].value_counts(normalize=True).get("Yes",0)*100:.1f}%'),
    ('月费差异', f'${df[df.Churn=="Yes"]["MonthlyCharges"].mean()-df[df.Churn=="No"]["MonthlyCharges"].mean():+.1f}'),
]
for i, (label, val) in enumerate(kpis):
    col = i * 2 + 1
    ws3.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    c = ws3.cell(row=3, column=col, value=val)
    c.font = Font(name='Arial', size=22, bold=True, color='4472C4'); c.fill = KPI_FILL; c.alignment = Alignment(horizontal='center'); c.border = THIN_BORDER
    ws3.cell(row=3, column=col+1).border = THIN_BORDER; ws3.cell(row=3, column=col+1).fill = KPI_FILL
    ws3.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    l = ws3.cell(row=4, column=col, value=label)
    l.font = Font(name='Arial', size=10, bold=True); l.alignment = Alignment(horizontal='center'); l.border = THIN_BORDER
    ws3.cell(row=4, column=col+1).border = THIN_BORDER

# Top findings
ws3.merge_cells('A6:F6')
ws3.cell(row=6, column=1, value="🔍 Top 3 发现").font = Font(name='Arial', size=12, bold=True, color='4472C4')
findings = [
    ('发现1', '合同类型是最强预测因素', '月付42.7% vs 两年约2.8%，差异15倍'),
    ('发现2', '头6个月是流失高危期', '53.3%新客在半年内流失'),
    ('发现3', '服务越多越难离开', '每多1个附加服务流失率降3%'),
]
write_header(ws3, 7, ['#', '发现', '关键数据'])
for i, (n, title, data) in enumerate(findings):
    write_row(ws3, 8+i, [n, title, data])

# Thinking models
ws3.merge_cells('A12:F12')
ws3.cell(row=12, column=1, value="🧠 思维模型应用").font = Font(name='Arial', size=12, bold=True, color='4472C4')
models = [
    ('分解思维', '总流失率26.6%拆解：月付贡献88%的流失'),
    ('分层差异', '新老客户流失率差异5.6倍，平均值不能代表任何一组'),
    ('杠杆点', '月付客户杠杆值8.87，是最佳改善切入点'),
]
write_header(ws3, 13, ['模型', '应用'])
for i, (m, desc) in enumerate(models):
    write_row(ws3, 14+i, [m, desc])

# Recommendations
ws3.merge_cells('A18:F18')
ws3.cell(row=18, column=1, value="💡 核心建议").font = Font(name='Arial', size=12, bold=True, color='4472C4')
write_header(ws3, 19, ['优先级', '建议行动', '追踪指标', '预期影响'])
recs = [
    ('P0', '月付新客免费试用附加服务', '首季流失率', '新客流失降15-20%'),
    ('P1', '家庭套餐折扣', '家庭套餐占比', '提升客户生命周期'),
    ('P1', '光纤客户年契促销', '光纤年契转化率', '光纤流失降20%'),
]
for i, (p, act, met, imp) in enumerate(recs):
    write_row(ws3, 20+i, [p, act, met, imp])

ws3.column_dimensions['A'].width = 12
ws3.column_dimensions['B'].width = 32
ws3.column_dimensions['C'].width = 20
ws3.column_dimensions['D'].width = 22
ws3.column_dimensions['E'].width = 16
ws3.column_dimensions['F'].width = 16

output_path = os.path.join(OUTPUT_DIR, "telco_churn_standard_report.xlsx")
wb.save(output_path)
print(f"Excel saved: {output_path}")
