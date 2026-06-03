"""
Retail Operations — Excel Deliverable
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter
import warnings; warnings.filterwarnings("ignore")

BASE = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/work/kaggle_datasets/downloads/store_retail_operations"
sales = pd.read_csv(f"{BASE}/sales data-set.csv", parse_dates=["Date"], date_format="%d/%m/%Y")
stores = pd.read_csv(f"{BASE}/stores data-set.csv")

sales_full = sales.merge(stores, on="Store", how="left")

wb = Workbook()
hfont = Font(bold=True, color="FFFFFF", size=11)
hfill = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
tfont = Font(bold=True, size=14, color="1F3864")
kfont = Font(bold=True, size=18, color="2F5496")
klfont = Font(size=9, color="666666")
bdr = Border(left=Side(style="thin"),right=Side(style="thin"),top=Side(style="thin"),bottom=Side(style="thin"))

def sh(ws, row, mc):
    for c in range(1, mc+1):
        cl = ws.cell(row=row, column=c)
        cl.font = hfont; cl.fill = hfill; cl.alignment = Alignment(horizontal="center"); cl.border = bdr
def sd(ws, sr, er, mc):
    for r in range(sr, er+1):
        for c in range(1, mc+1):
            ws.cell(row=r, column=c).border = bdr

# Sheet 1
ws1 = wb.active; ws1.title = "整体概览"
ws1.merge_cells("A1:D1")
ws1["A1"] = "Store Retail — 关键指标"; ws1["A1"].font = tfont
kpis = [("总销售额", f"${sales['Weekly_Sales'].sum()/1e9:.2f}B"), ("平均周销", f"${sales['Weekly_Sales'].mean():,.0f}"),
        ("门店数", f"{stores['Store'].nunique()}"), ("部门数", f"{sales['Dept'].nunique()}"),
        ("A 型门店均销", f"${sales_full[sales_full['Type']=='A']['Weekly_Sales'].mean():,.0f}"),
        ("B 型门店均销", f"${sales_full[sales_full['Type']=='B']['Weekly_Sales'].mean():,.0f}")]
for i,(l,v) in enumerate(kpis):
    ws1.cell(row=3+i, column=1, value=l).font = Font(bold=True)
    ws1.cell(row=3+i, column=2, value=v).font = kfont

# Sheet 2: Store Type
ws2 = wb.create_sheet("门店类型")
ws2.merge_cells("A1:D1"); ws2["A1"] = "门店类型表现"; ws2["A1"].font = tfont
st = sales_full.groupby("Type").agg(门店数=("Store","nunique"),总销售额=("Weekly_Sales","sum"),平均周销=("Weekly_Sales","mean")).round(2).reset_index()
for r in dataframe_to_rows(st, index=False, header=True): ws2.append(r)
sh(ws2, 3, len(st.columns)); sd(ws2, 4, 3+len(st), len(st.columns))

# Sheet 3: Top/Bottom Depts
ws3 = wb.create_sheet("部门表现")
ws3.merge_cells("A1:D1"); ws3["A1"] = "部门 Top 15"; ws3["A1"].font = tfont
dept = sales.groupby("Dept").agg(总销售额=("Weekly_Sales","sum"),平均周销=("Weekly_Sales","mean")).sort_values("总销售额", ascending=False).head(15).reset_index()
for r in dataframe_to_rows(dept, index=False, header=True): ws3.append(r)
sh(ws3, 3, len(dept.columns)); sd(ws3, 4, 3+len(dept), len(dept.columns))

# Sheet 4: Monthly
ws4 = wb.create_sheet("月度趋势")
ws4.merge_cells("A1:C1"); ws4["A1"] = "月度销售趋势"; ws4["A1"].font = tfont
m = sales.copy(); m["YearMonth"] = m["Date"].dt.to_period("M").astype(str)
ma = m.groupby("YearMonth").agg(销售额=("Weekly_Sales","sum"),平均周销=("Weekly_Sales","mean")).reset_index()
for r in dataframe_to_rows(ma, index=False, header=True): ws4.append(r)
sh(ws4, 3, len(ma.columns)); sd(ws4, 4, 3+len(ma), len(ma.columns))

# Sheet 5: Holiday effect
ws5 = wb.create_sheet("节假日效应")
ws5.merge_cells("A1:D1"); ws5["A1"] = "节假日 vs 非节假日"; ws5["A1"].font = tfont
hd = sales.groupby("IsHoliday").agg(销售额=("Weekly_Sales","sum"),平均周销=("Weekly_Sales","mean"),记录数=("Weekly_Sales","count")).reset_index()
for r in dataframe_to_rows(hd, index=False, header=True): ws5.append(r)
sh(ws5, 3, len(hd.columns)); sd(ws5, 4, 3+len(hd), len(hd.columns))

# Column widths
for ws in [ws1, ws2, ws3, ws4, ws5]:
    for ci in range(1, ws.max_column+1):
        vals = [len(str(ws.cell(row=ri, column=ci).value or "")) for ri in range(1, ws.max_row+1)]
        ws.column_dimensions[get_column_letter(ci)].width = min(max(vals)+4, 50)

out = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output/retail_analysis.xlsx"
wb.save(out); print(f"Excel: {out}")
