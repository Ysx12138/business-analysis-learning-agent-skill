"""
Generate Excel pivot tables and dashboard for Telco Churn Analysis
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.series import DataPoint
from openpyxl.utils import get_column_letter
import os

DATA_PATH = "/Users/sx/数据分析/telco_customer_churn/原始数据/telco_churn.csv"
OUTPUT_DIR = "/Users/sx/Desktop/AI-Data-Analysis-Learning-Skill/github/business-analysis-learning-agent-skill/output"
os.makedirs(OUTPUT_DIR, exist_ok=True)

df = pd.read_csv(DATA_PATH)
df['TotalCharges'] = pd.to_numeric(df['TotalCharges'], errors='coerce')
df = df.dropna(subset=['TotalCharges'])
df = df[df['tenure'] > 0]

# Derived fields
df['tenure_group'] = pd.cut(df['tenure'],
    bins=[-1, 6, 12, 24, 48, 72],
    labels=['0-6月', '7-12月', '13-24月', '25-48月', '49-72月'])
service_fields = ['OnlineSecurity', 'OnlineBackup', 'DeviceProtection',
                  'TechSupport', 'StreamingTV', 'StreamingMovies']
for f in service_fields:
    df[f+'_bin'] = df[f].apply(lambda x: 0 if x=='No' else (1 if x=='Yes' else 0))
df['service_count'] = sum(df[f+'_bin'] for f in service_fields)
df['has_family'] = (df['Partner']=='Yes') | (df['Dependents']=='Yes')
df['has_security_and_support'] = (df['OnlineSecurity']=='Yes') & (df['TechSupport']=='Yes')

BASE_RATE = df['Churn'].value_counts(normalize=True)['Yes']

# ============ STYLES ============
TITLE_FONT = Font(name='Arial', size=14, bold=True, color='FFFFFF')
HEADER_FONT = Font(name='Arial', size=11, bold=True, color='FFFFFF')
DATA_FONT = Font(name='Arial', size=10)
PCT_FONT = Font(name='Arial', size=10, color='CC0000', bold=True)
HEADER_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
ALT_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
HIGHLIGHT_FILL = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
DASHBOARD_FILL = PatternFill(start_color='2F5496', end_color='2F5496', fill_type='solid')
KPI_FILL = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'))

def write_header_row(ws, row, headers, start_col=1):
    for i, h in enumerate(headers):
        cell = ws.cell(row=row, column=start_col+i, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER

def write_data_row(ws, row, values, start_col=1, pct_cols=None, highlight=False):
    if pct_cols is None:
        pct_cols = []
    for i, v in enumerate(values):
        cell = ws.cell(row=row, column=start_col+i, value=v)
        cell.font = DATA_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = THIN_BORDER
        if i in pct_cols and isinstance(v, (int, float)):
            cell.number_format = '0.0%'
        if highlight:
            cell.fill = ALT_FILL

# ============ CREATE WORKBOOK ============
wb = Workbook()

# ========================
# SHEET 1: 整体对比
# ========================
ws1 = wb.active
ws1.title = "整体对比"
ws1.sheet_properties.tabColor = '2F5496'

# Title
ws1.merge_cells('A1:E1')
title_cell = ws1.cell(row=1, column=1, value="📊 Telco Churn - 整体对比")
title_cell.font = TITLE_FONT
title_cell.fill = DASHBOARD_FILL
title_cell.alignment = Alignment(horizontal='center', vertical='center')

# Overview table
row = 3
write_header_row(ws1, row, ['指标', '总体', '未流失 (No)', '流失 (Yes)', '差异倍数'])
metrics = [
    ('客户数', len(df), df[df.Churn=='No'].shape[0], df[df.Churn=='Yes'].shape[0], '-'),
    ('流失率', BASE_RATE, '-', '-', '-'),
    ('平均在网月数', df['tenure'].mean(), df[df.Churn=='No']['tenure'].mean(),
     df[df.Churn=='Yes']['tenure'].mean(), df[df.Churn=='No']['tenure'].mean()/max(df[df.Churn=='Yes']['tenure'].mean(),0.01)),
    ('平均月费', df['MonthlyCharges'].mean(), df[df.Churn=='No']['MonthlyCharges'].mean(),
     df[df.Churn=='Yes']['MonthlyCharges'].mean(), '-'),
    ('平均总费用', df['TotalCharges'].mean(), df[df.Churn=='No']['TotalCharges'].mean(),
     df[df.Churn=='Yes']['TotalCharges'].mean(), '-'),
    ('平均附加服务数', df['service_count'].mean(), df[df.Churn=='No']['service_count'].mean(),
     df[df.Churn=='Yes']['service_count'].mean(), '-'),
    ('有家庭比例', df['has_family'].mean(), df[df.Churn=='No']['has_family'].mean(),
     df[df.Churn=='Yes']['has_family'].mean(), '-'),
]

for i, (label, total, no, yes, ratio) in enumerate(metrics):
    r = row + 1 + i
    pct_cols = set()
    vals = [label, total, no, yes, ratio]
    if label == '流失率':
        pct_cols = {1}
    write_data_row(ws1, r, vals, pct_cols=pct_cols, highlight=(i%2==1))

# Column widths
ws1.column_dimensions['A'].width = 18
ws1.column_dimensions['B'].width = 14
ws1.column_dimensions['C'].width = 16
ws1.column_dimensions['D'].width = 14
ws1.column_dimensions['E'].width = 12

# ========================
# SHEET 2: 按合同类型
# ========================
ws2 = wb.create_sheet("按合同类型")
ws2.sheet_properties.tabColor = '4472C4'

ws2.merge_cells('A1:F1')
t2 = ws2.cell(row=1, column=1, value="📋 按合同类型分析")
t2.font = TITLE_FONT; t2.fill = DASHBOARD_FILL; t2.alignment = Alignment(horizontal='center')

write_header_row(ws2, 3, ['合同类型', '客户数', '占比', '未流失', '流失', '流失率'])
for i, (ctype, grp) in enumerate(df.groupby('Contract')):
    r = 4 + i
    total = len(grp)
    churn_yes = grp[grp.Churn=='Yes'].shape[0]
    churn_no = grp[grp.Churn=='No'].shape[0]
    rate = churn_yes / total
    write_data_row(ws2, r, [ctype, total, total/len(df), churn_no, churn_yes, rate],
                   pct_cols={2, 5}, highlight=(i%2==1))

# Chart: churn rate by contract
chart1 = BarChart()
chart1.type = "col"
chart1.title = "各合同类型流失率"
chart1.y_axis.title = "流失率"
chart1.style = 10
data_ref = Reference(ws2, min_col=6, min_row=3, max_row=6)
cats_ref = Reference(ws2, min_col=1, min_row=4, max_row=6)
chart1.add_data(data_ref, titles_from_data=True)
chart1.set_categories(cats_ref)
chart1.width = 16
chart1.height = 10
ws2.add_chart(chart1, "A8")

ws2.column_dimensions['A'].width = 18
ws2.column_dimensions['B'].width = 10
for c in 'CDEF':
    ws2.column_dimensions[c].width = 12

# ========================
# SHEET 3: 按在网时长
# ========================
ws3 = wb.create_sheet("按在网时长")
ws3.sheet_properties.tabColor = '5B9BD5'

ws3.merge_cells('A1:F1')
t3 = ws3.cell(row=1, column=1, value="📋 按在网时长分组分析")
t3.font = TITLE_FONT; t3.fill = DASHBOARD_FILL; t3.alignment = Alignment(horizontal='center')

write_header_row(ws3, 3, ['在网时长', '客户数', '占比', '未流失', '流失', '流失率'])
tenure_order = ['0-6月', '7-12月', '13-24月', '25-48月', '49-72月']
for i, tg in enumerate(tenure_order):
    r = 4 + i
    grp = df[df['tenure_group'] == tg]
    total = len(grp)
    churn_yes = grp[grp.Churn=='Yes'].shape[0]
    churn_no = grp[grp.Churn=='No'].shape[0]
    rate = churn_yes / total
    write_data_row(ws3, r, [tg, total, total/len(df), churn_no, churn_yes, rate],
                   pct_cols={2, 5}, highlight=(i%2==1))

chart3 = BarChart()
chart3.type = "col"
chart3.title = "在网时长分组流失率"
chart3.style = 10
data_ref = Reference(ws3, min_col=6, min_row=3, max_row=8)
cats_ref = Reference(ws3, min_col=1, min_row=4, max_row=8)
chart3.add_data(data_ref, titles_from_data=True)
chart3.set_categories(cats_ref)
chart3.width = 16
chart3.height = 10
ws3.add_chart(chart3, "A10")

ws3.column_dimensions['A'].width = 14
ws3.column_dimensions['B'].width = 10
for c in 'CDEF':
    ws3.column_dimensions[c].width = 12

# ========================
# SHEET 4: 按服务类别
# ========================
ws4 = wb.create_sheet("按服务类别")
ws4.sheet_properties.tabColor = '8DB4E2'

ws4.merge_cells('A1:F1')
t4 = ws4.cell(row=1, column=1, value="📋 按服务类别分析")
t4.font = TITLE_FONT; t4.fill = DASHBOARD_FILL; t4.alignment = Alignment(horizontal='center')

write_header_row(ws4, 3, ['服务维度', '类别', '客户数', '占比', '流失数', '流失率'])

svc_dims = [
    ('互联网服务', 'InternetService'),
    ('支付方式', 'PaymentMethod'),
    ('在线安全', 'OnlineSecurity'),
    ('技术支持', 'TechSupport'),
    ('在线备份', 'OnlineBackup'),
    ('设备保护', 'DeviceProtection'),
    ('电子账单', 'PaperlessBilling'),
    ('家庭绑定', 'has_family'),
    ('安全+支持套餐', 'has_security_and_support'),
]
r = 4
for label, col in svc_dims:
    grps = df.groupby(col)
    first = True
    for cat, grp in grps:
        total = len(grp)
        churn_yes = grp[grp.Churn=='Yes'].shape[0]
        rate = churn_yes / total
        display_label = label if first else ''
        write_data_row(ws4, r, [display_label, cat, total, total/len(df), churn_yes, rate],
                       pct_cols={3, 5}, highlight=(r%2==0))
        first = False
        r += 1

ws4.column_dimensions['A'].width = 18
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 10
ws4.column_dimensions['D'].width = 10
ws4.column_dimensions['E'].width = 10
ws4.column_dimensions['F'].width = 10

# ========================
# SHEET 5: 交叉分析
# ========================
ws5 = wb.create_sheet("交叉分析")
ws5.sheet_properties.tabColor = 'A8D08D'

ws5.merge_cells('A1:G1')
t5 = ws5.cell(row=1, column=1, value="📋 交叉分析 - 合同类型 × 在网时长")
t5.font = TITLE_FONT; t5.fill = DASHBOARD_FILL; t5.alignment = Alignment(horizontal='center')

write_header_row(ws5, 3, ['合同类型', '0-6月', '7-12月', '13-24月', '25-48月', '49-72月', '整体'])
contracts_order = ['Month-to-month', 'One year', 'Two year']
for i, ct in enumerate(contracts_order):
    r = 4 + i
    row_data = [ct]
    for tg in tenure_order:
        sub = df[(df['Contract']==ct) & (df['tenure_group']==tg)]
        if len(sub) > 0:
            rate = sub[sub.Churn=='Yes'].shape[0] / len(sub)
            row_data.append(rate)
        else:
            row_data.append(0)
    row_data.append(contract_churn := df[df['Contract']==ct]['Churn'].value_counts(normalize=True).get('Yes', 0))
    write_data_row(ws5, r, row_data, pct_cols={1,2,3,4,5,6}, highlight=(i%2==1))

# Chart: stacked bar
chart5 = BarChart()
chart5.type = "col"
chart5.grouping = "clustered"
chart5.title = "合同类型 × 在网时长 流失率"
chart5.style = 10
data_ref = Reference(ws5, min_col=2, max_col=7, min_row=3, max_row=6)
cats_ref = Reference(ws5, min_col=1, min_row=4, max_row=6)
chart5.add_data(data_ref, titles_from_data=True)
chart5.set_categories(cats_ref)
chart5.width = 20
chart5.height = 12
ws5.add_chart(chart5, "A8")

ws5.column_dimensions['A'].width = 18
for c in 'BCDEFG':
    ws5.column_dimensions[c].width = 14

# ========================
# SHEET 6: 分析看板
# ========================
ws6 = wb.create_sheet("分析看板")
ws6.sheet_properties.tabColor = 'FFC000'

ws6.merge_cells('A1:F1')
t6 = ws6.cell(row=1, column=1, value="📊 Telco Churn 分析看板")
t6.font = Font(name='Arial', size=16, bold=True, color='FFFFFF')
t6.fill = DASHBOARD_FILL
t6.alignment = Alignment(horizontal='center', vertical='center')
ws6.row_dimensions[1].height = 30

# KPI row
kpi_data = [
    ('整体流失率', f'{BASE_RATE*100:.1f}%', '26.5% 是电信行业偏高水平，需重点关注', 'C4'),
    ('月付客户占比', f'{df[df["Contract"]=="Month-to-month"].shape[0]/len(df)*100:.1f}%', '超过一半客户是月付，缺乏长期绑定', 'C5'),
    ('前6月流失率', f'{df[df["tenure_group"]=="0-6月"]["Churn"].value_counts(normalize=True).get("Yes",0)*100:.1f}%', '每2个新客有1个在头半年流失', 'C6'),
    ('月费差异', f'${df[df.Churn=="Yes"]["MonthlyCharges"].mean()-df[df.Churn=="No"]["MonthlyCharges"].mean():+.1f}', '流失客户月费平均高$13.2，可能价格敏感', 'C7'),
]
for i, (label, value, note, _) in enumerate(kpi_data):
    col = 2 * i + 1
    # KPI box
    ws6.merge_cells(start_row=3, start_column=col, end_row=3, end_column=col+1)
    kpi_cell = ws6.cell(row=3, column=col, value=value)
    kpi_cell.font = Font(name='Arial', size=24, bold=True, color='2F5496')
    kpi_cell.fill = KPI_FILL
    kpi_cell.alignment = Alignment(horizontal='center', vertical='center')
    kpi_cell.border = THIN_BORDER
    ws6.cell(row=3, column=col+1).border = THIN_BORDER
    ws6.cell(row=3, column=col+1).fill = KPI_FILL
    # Label
    ws6.merge_cells(start_row=4, start_column=col, end_row=4, end_column=col+1)
    l_cell = ws6.cell(row=4, column=col, value=label)
    l_cell.font = Font(name='Arial', size=11, bold=True, color='2F5496')
    l_cell.alignment = Alignment(horizontal='center')
    l_cell.border = THIN_BORDER
    ws6.cell(row=4, column=col+1).border = THIN_BORDER
    # Note
    ws6.merge_cells(start_row=5, start_column=col, end_row=5, end_column=col+1)
    n_cell = ws6.cell(row=5, column=col, value=note)
    n_cell.font = Font(name='Arial', size=9, color='666666')
    n_cell.alignment = Alignment(horizontal='center', wrap_text=True)
    n_cell.border = THIN_BORDER
    ws6.cell(row=5, column=col+1).border = THIN_BORDER

# Thinking Models Section
ws6.merge_cells('A7:F7')
tm_title = ws6.cell(row=7, column=1, value="🧠 5种思维模型提炼")
tm_title.font = Font(name='Arial', size=13, bold=True, color='2F5496')
tm_title.alignment = Alignment(horizontal='left')

thinking_models = [
    ('A. 分解思维', '将总流失率 26.6% 分解到各维度：\n• 月付客户贡献了总流失的 23.5%（占55%客户，流失率42.7%）\n• 前6个月新客贡献了总流失的 16.2%（占26%客户，流失率53.3%）\n→ 改善优先级：月付 > 新客 > 光纤客户'),
    ('B. 分层差异', '在网时长组之间流失率差异高达 5.6 倍：\n• 0-6月: 53.3% | 7-12月: 35.9% | 13-24月: 28.7% | 25-48月: 20.4% | 49-72月: 9.5%\n→ 不要只看平均值（26.6%），新老客户行为完全不同'),
    ('C. 代理推断', '"服务粘性"无法直接测量，用以下代理变量推断：\n• 附加服务数量：0个=21.4%流失 → 6个=5.3%流失（每多1个服务，流失率降约3%）\n• 在线安全+技术支持双重用户流失率仅 12.0%\n→ 附加服务数量是客户粘性的有效代理指标'),
    ('D. 约束vs偏好', '光纤客户流失率41.9% vs DSL客户19.0%：\n• 约束假说：光纤网络质量或价格有问题 → 需要修网络/调价\n• 偏好假说：68.7%光纤客户选月付（DSL仅50.6%）→ 客户群本身更短期导向\n• 结论：两种因素并存，但"偏好"被数据更充分支持\n→ 行动应侧重合约设计而非网络投资'),
    ('E. 杠杆点', '寻找高占比 × 高改进潜力的细分：\n• 月付客户：占比55.1%，流失率42.7% → 杠杆值 ≈ 55% × (42.7%-26.6%) = 8.9\n• 前6月新客：占比26.3%，流失率53.3% → 杠杆值 ≈ 26% × (53.3%-26.6%) = 7.0\n• 最高杠杆：改善月付客户的入网体验，特别是前6个月'),
]

for i, (title, content) in enumerate(thinking_models):
    r = 8 + i * 2
    ws6.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    cell = ws6.cell(row=r, column=1, value=title)
    cell.font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = Alignment(horizontal='left', vertical='center')

    ws6.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=6)
    cell_c = ws6.cell(row=r+1, column=1, value=content)
    cell_c.font = Font(name='Arial', size=10)
    cell_c.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
    cell_c.fill = PatternFill(start_color='F2F7FB', end_color='F2F7FB', fill_type='solid')
    ws6.row_dimensions[r+1].height = 60

# Recommendations Section
rec_row = 19
ws6.merge_cells(f'A{rec_row}:F{rec_row}')
rec_title = ws6.cell(row=rec_row, column=1, value="💡 核心建议")
rec_title.font = Font(name='Arial', size=13, bold=True, color='2F5496')
rec_title.alignment = Alignment(horizontal='left')

write_header_row(ws6, rec_row+1, ['优先级', '建议行动', '预期影响', '追踪指标', '投入成本'])
recommendations = [
    ('P0', '月付新客: 前3个月免费试用在线安全+技术支持', '预计降低新客流失率15-20%', '首季流失率、安全包渗透率', '低（免费试用）'),
    ('P1', '推出家庭套餐折扣（绑定配偶/子女）', '有家庭客户流失率仅14.2%（vs 34.2%）', '家庭套餐占比、流失率', '中（套餐设计）'),
    ('P1', '光纤客户: 提供6个月合约折扣而非月付', '光纤客户月付比例68.7%，转合约可降流失', '光纤合约转化率', '中（收入延迟）'),
    ('P2', '流失预警: 对月付+前6月+高月费客户主动外呼', '精准覆盖最高危群体', '干预覆盖率、挽回率', '高（人力成本）'),
]
for i, (pri, action, impact, metric, cost) in enumerate(recommendations):
    r = rec_row + 2 + i
    write_data_row(ws6, r, [pri, action, impact, metric, cost], highlight=(i%2==1))
    if pri == 'P0':
        for c in range(1, 6):
            ws6.cell(row=r, column=c).fill = HIGHLIGHT_FILL

# Column widths for dashboard
ws6.column_dimensions['A'].width = 10
ws6.column_dimensions['B'].width = 40
ws6.column_dimensions['C'].width = 26
ws6.column_dimensions['D'].width = 22
ws6.column_dimensions['E'].width = 16
ws6.column_dimensions['F'].width = 16

# Save
output_path = os.path.join(OUTPUT_DIR, "telco_churn_audit_report.xlsx")
wb.save(output_path)
print(f"Excel 已保存: {output_path}")
