"""
Generic Excel Renderer

Consumes a report dict (result_schema) and produces an .xlsx file.
No dataset-specific logic here — everything is driven by the report structure.
"""
import os
import warnings
from openpyxl import Workbook
from openpyxl.chart import BarChart, PieChart, Reference
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

# ── Style constants ──
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
HEADER_FILL = PatternFill("solid", fgColor="1B365D")
ACCENT_FILL = PatternFill("solid", fgColor="EEF2F7")
DASH_FILL = PatternFill("solid", fgColor="F5F4ED")
TITLE_FONT = Font(bold=True, size=14, color="1B365D")
SECTION_FONT = Font(bold=True, size=11, color="1B365D")
BODY_FONT = Font(size=10, color="333333")
KPI_VALUE_FONT = Font(name="Helvetica", size=18, bold=True, color="1B365D")
KPI_LABEL_FONT = Font(size=8.5, color="666666")
THIN_BORDER = Border(
    left=Side(style="thin", color="D0D0D0"),
    right=Side(style="thin", color="D0D0D0"),
    top=Side(style="thin", color="D0D0D0"),
    bottom=Side(style="thin", color="D0D0D0"),
)


def _style_header(ws, row=1, max_col=None):
    if max_col is None:
        max_col = ws.max_column
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _auto_width(ws):
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                val = str(cell.value or "")
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 35)


def _add_table(ws, headers, data, start_row=1):
    """Write a table with styled header. Returns the next empty row."""
    for ci, h in enumerate(headers, 1):
        ws.cell(row=start_row, column=ci, value=h)
    _style_header(ws, row=start_row, max_col=len(headers))
    for ri, row_data in enumerate(data, start_row + 1):
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.font = Font(size=9.5, color="333333")
    return start_row + len(data) + 1


def _add_chart(ws, data_ref, cat_ref, anchor, title="", chart_type="bar"):
    """Add a chart anchored at a cell."""
    if chart_type == "bar":
        chart = BarChart()
        chart.type = "col"
    elif chart_type == "pie":
        chart = PieChart()
    else:
        chart = BarChart()
    chart.title = title
    chart.style = 2
    chart.width = 16
    chart.height = 10
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cat_ref)
    ws.add_chart(chart, anchor)


def _safe_sheet_name(name: str, max_len: int = 25) -> str:
    """Truncate sheet name to leave room for OpenPyXL dedup suffix."""
    cleaned = "".join(c for c in name if c.isprintable() and c not in r"\/?*[]:")
    if len(cleaned) > max_len:
        cleaned = cleaned[:max_len]
    return cleaned


def render(report: dict, output_path: str) -> str:
    """
    Main entry point. Takes a report dict (from result_schema) and writes an .xlsx file.
    Returns the output path.
    """
    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Title is more than 31 characters")
        wb = Workbook()
        overview = report.get("dataset_overview", {})
        tables = report.get("tables", [])
        findings = report.get("key_findings", [])
        models = report.get("thinking_models", [])
        recs = report.get("recommendations", [])
        metrics = report.get("metric_glossary", [])
        semantics = report.get("field_semantics", [])
        quality = report.get("data_quality", {})

    # ── Sheet 1: Dataset Overview ──
    ws1 = wb.active
    ws1.title = "数据概览"
    ws1.cell(row=1, column=1, value="数据集概览").font = TITLE_FONT

    meta_rows = [
        ("文件名", overview.get("file_name", "")),
        ("记录数", str(overview.get("rows", 0))),
        ("字段数", str(overview.get("columns", 0))),
        ("缺失值", "有" if overview.get("has_missing") else "无"),
        ("重复行", str(overview.get("duplicate_rows", 0))),
        ("时间范围", overview.get("time_range", "")),
    ]
    for i, (k, v) in enumerate(meta_rows, 3):
        ws1.cell(row=i, column=1, value=k).font = Font(bold=True, size=10)
        ws1.cell(row=i, column=2, value=v).font = Font(size=10)

    # Fields table
    fields = overview.get("fields", [])
    if fields:
        r = 11
        ws1.cell(row=r, column=1, value="字段列表").font = SECTION_FONT
        r += 1
        ftypes = overview.get("field_types", {})
        fheaders = ["字段名", "类型"]
        _add_table(ws1, fheaders, [(f, ftypes.get(f, "")) for f in fields], start_row=r)

    # Field semantics
    if semantics:
        r = r + len(fields) + 4
        ws1.cell(row=r, column=1, value="字段语义").font = SECTION_FONT
        r += 1
        sheaders = ["字段", "推断含义", "置信度"]
        sdata = [(s["field"], s["inferred_meaning"], s["confidence"]) for s in semantics]
        _add_table(ws1, sheaders, sdata, start_row=r)

    _auto_width(ws1)

    # ── Sheets: Data Tables ──
    for ti, tbl in enumerate(tables):
        ws = wb.create_sheet(_safe_sheet_name(tbl.get("name", f"Sheet{ti+2}")))
        _add_table(ws, tbl.get("headers", []), tbl.get("data", []))
        _auto_width(ws)

        # Add bar chart if table has numeric columns and enough rows
        data_rows = tbl.get("data", [])
        headers = tbl.get("headers", [])
        if len(data_rows) >= 2 and len(headers) >= 2:
            # Find first numeric column beyond col 1
            numeric_cols = []
            for ci in range(1, len(headers)):
                col_vals = [row[ci] for row in data_rows if ci < len(row)]
                if col_vals and all(isinstance(v, (int, float)) for v in col_vals):
                    numeric_cols.append(ci + 1)
            if numeric_cols:
                chart_start_row = len(data_rows) + 4
                data_ref = Reference(ws, min_col=numeric_cols[0], min_row=1,
                                     max_row=len(data_rows) + 1, max_col=numeric_cols[0])
                cat_ref = Reference(ws, min_col=1, min_row=2, max_row=len(data_rows) + 1)
                anchor_col = len(headers) + 3
                anchor_cell = f"{get_column_letter(anchor_col)}{chart_start_row - 3}"
                _add_chart(ws, data_ref, cat_ref, anchor_cell,
                           title=f"{headers[numeric_cols[0]-1]} 分布")

    # ── Dashboard Sheet ──
    ws_dash = wb.create_sheet("分析看板")
    ws_dash.merge_cells("A1:H1")
    ws_dash.cell(row=1, column=1, value=report.get("title", "分析看板")).font = TITLE_FONT
    ws_dash.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    row = 3

    # KPI row
    if metrics:
        ws_dash.cell(row=row, column=1, value="关键指标").font = SECTION_FONT
        row += 1
        # Show first 6 metrics as KPI boxes
        kpi_metrics = metrics[:6]
        for i, m in enumerate(kpi_metrics):
            col = i + 1
            cell_val = ws_dash.cell(row=row, column=col, value=str(m.get("current_value", "")))
            cell_val.font = KPI_VALUE_FONT
            cell_val.alignment = Alignment(horizontal="center")
            cell_label = ws_dash.cell(row=row + 1, column=col, value=m.get("name", ""))
            cell_label.font = KPI_LABEL_FONT
            cell_label.alignment = Alignment(horizontal="center")
        row += 3

    # Key Findings
    if findings:
        ws_dash.cell(row=row, column=1, value="核心发现").font = SECTION_FONT
        row += 1
        for fi, f in enumerate(findings):
            ws_dash.cell(row=row, column=1, value=f"{fi+1}. {f['title']}").font = Font(bold=True, size=10, color="1B365D")
            ws_dash.cell(row=row, column=2, value=f["data_evidence"]).font = BODY_FONT
            row += 1
        row += 1

    # Thinking Models
    if models:
        ws_dash.cell(row=row, column=1, value="思维模型").font = SECTION_FONT
        row += 1
        for m in models:
            ws_dash.cell(row=row, column=1, value=m["model_name"]).font = Font(bold=True, size=10)
            ws_dash.cell(row=row, column=2, value=m["takeaway"]).font = BODY_FONT
            row += 1
        row += 1

    # Recommendations
    if recs:
        ws_dash.cell(row=row, column=1, value="行动建议").font = SECTION_FONT
        row += 1
        rheaders = ["建议", "行动", "追踪指标"]
        rdata = [(r["title"], r["action"], r["metric_to_track"]) for r in recs]
        row = _add_table(ws_dash, rheaders, rdata, start_row=row)

    # ── Learning Guide ──
    beginner = report.get("beginner_notes", {})
    if beginner:
        row += 1
        ws_dash.cell(row=row, column=1, value="学习要点").font = SECTION_FONT
        row += 1

        concepts = beginner.get("concepts_learned", [])
        methods_learned = beginner.get("methods_learned", [])
        metrics_l = beginner.get("metrics_learned", [])
        patterns = beginner.get("thinking_patterns", [])

        if concepts:
            ws_dash.cell(row=row, column=1, value="学到的概念：").font = Font(bold=True, size=10)
            ws_dash.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws_dash.cell(row=row, column=2, value="、".join(concepts)).font = BODY_FONT
            row += 1
        if methods_learned:
            ws_dash.cell(row=row, column=1, value="学到的方法：").font = Font(bold=True, size=10)
            ws_dash.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws_dash.cell(row=row, column=2, value="、".join(methods_learned)).font = BODY_FONT
            row += 1
        if metrics_l:
            ws_dash.cell(row=row, column=1, value="学到的指标：").font = Font(bold=True, size=10)
            ws_dash.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws_dash.cell(row=row, column=2, value="、".join(metrics_l)).font = BODY_FONT
            row += 1
        if patterns:
            ws_dash.cell(row=row, column=1, value="应用的思维模式：").font = Font(bold=True, size=10)
            ws_dash.merge_cells(start_row=row, start_column=2, end_row=row, end_column=6)
            ws_dash.cell(row=row, column=2, value="、".join(patterns)).font = BODY_FONT
            row += 1

        next_steps = beginner.get("next_time_steps", [])
        if next_steps:
            row += 1
            ws_dash.cell(row=row, column=1, value="下次怎么做").font = Font(bold=True, size=10, color="1B365D")
            row += 1
            for ns in next_steps:
                ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws_dash.cell(row=row, column=1, value=f"• {ns}").font = BODY_FONT
                row += 1

        caveats = beginner.get("conclusions_not_to_overinterpret", [])
        if caveats:
            row += 1
            ws_dash.cell(row=row, column=1, value="不要过度解读").font = Font(bold=True, size=10, color="C00000")
            row += 1
            for cv in caveats:
                ws_dash.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
                ws_dash.cell(row=row, column=1, value=f"• {cv}").font = Font(size=9.5, color="666666")
                row += 1

    # Column widths for dashboard
    for ci in range(1, 9):
        ws_dash.column_dimensions[get_column_letter(ci)].width = 22

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", message="Title is more than 31 characters")
        wb.save(output_path)
    return output_path
